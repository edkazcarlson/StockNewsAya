from textblob import TextBlob
from datetime import datetime
from datetime import timedelta
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl import worksheet
from openpyxl.utils.cell import get_column_letter
import os.path



class excelManager:
	def __init__(self, metaData, stockDataTags):
		if os.path.exists('res/dataSheet.xlsx') :
			self.wb = load_workbook('res/dataSheet.xlsx')
			self.stockPage = self.wb["Stock Page"]
			self.forexPage = self.wb["Forex Page"]
			self.wb.save('res/dataSheet.xlsx')
		else: 
			self.wb = Workbook()
			self.firstTimeInit()
			self.wb.save('res/dataSheet.xlsx')
		self.metaData = metaData
		self.stockDataTags = stockDataTags
		self.newTag(metaData)

	#ran if the spreadsheet needs initial values and pages
	def firstTimeInit(self):
		self.stockPage = self.wb.active
		self.stockPage.title = "Stock Page"
		self.stockPage.cell(row=2, column=1, value='Average news score for the day')
		self.forexPage = self.wb.create_sheet(title = "Forex Page")
		self.forexPage.cell(row=2, column=1, value='Average news score for the day')
	
	#reads back the most recent line 
	def readMostRecentLine(self):
		print()
		
	def save(self):
		self.wb.save('res/dataSheet.xlsx')
		
	def writeNewLine(self, todayDataDict):
		print("\n\n")
		todaysColumn = self.stockPage.max_column + 1
		curRow = 1
		self.stockPage.cell(row = curRow, column = todaysColumn, value = todayDataDict['todayDateStr'])
		curRow += 1
		self.stockPage.cell(row = curRow, column = todaysColumn, value = todayDataDict['newsScore'])
		curRow += 1
		for x in todayDataDict['Stocks']:
			for y in todayDataDict['Stocks'][x]:
				self.stockPage.cell(row = curRow, column = todaysColumn, value = todayDataDict['Stocks'][x][y])
				curRow += 1
				
				
		todaysColumn = self.forexPage.max_column + 1
		curRow = 1
		self.forexPage.cell(row = curRow, column = todaysColumn, value = todayDataDict['todayDateStr'])
		curRow += 1
		self.forexPage.cell(row = curRow, column = todaysColumn, value = todayDataDict['newsScore'])
		curRow += 1
		for x in todayDataDict['Forex']:
			for y in todayDataDict['Forex'][x]:
				self.forexPage.cell(row = curRow, column = todaysColumn, value = todayDataDict['Forex'][x][y])
				curRow += 1
		
	def readXMostRecentLines(x):
		print()
		
	#returns a list of tags currently being used
	def getTags(self):
		return metaData
	
	#call when there's a new tag in order to add to the spreadsheet 
	def newTag(self, metaData):
		curRow = 3
		self.metaData = metaData
		for x in metaData['stocksWatched']:
			for y in self.stockDataTags:
				print(self.stockPage['A' + str(curRow)].value)
				if self.stockPage['A' + str(curRow)].value != x + ' ' + y: #if it changed, needs to move everything down 1
					self.stockPage.move_range('A' + str(curRow) + ':' + get_column_letter(self.stockPage.max_column) + str(self.stockPage.max_row), rows = 1, cols = 0)
				self.stockPage.cell(row=curRow, column=1, value=x + ' ' + y)
				curRow += 1
		
		curRow = 3
		for x in metaData['forexWatched']:
			print(x)
			for y in metaData['forexWatched'][x]:
				if self.forexPage['A' + str(curRow)].value != x + '->' + y: #if it changed, needs to move everything down 1
					self.forexPage.move_range('A' + str(curRow) + ':' + get_column_letter(self.forexPage.max_column) + str(self.forexPage.max_row), rows = 1, cols = 0)
				self.forexPage.cell(row=curRow, column=1, value=x + '->' + y)
				curRow += 1
			
	#call when there's a tag removed
	def removeTag(self, metaData):
		curRow = 3
		for x in metaData['stocksWatched']:
			for y in self.stockDataTags:
				print(self.stockPage['A' + str(curRow)].value)
				if self.stockPage['A' + str(curRow)].value != x + ' ' + y: #if it changed, needs to move everything up 1
					self.stockPage.move_range('A' + str(curRow) + ':' + get_column_letter(self.stockPage.max_column) + str(self.stockPage.max_row), rows = -1, cols = 0)
				else:
					self.stockPage.cell(row=curRow, column=1, value=x + ' ' + y)
					curRow += 1
				
		curRow = 3
		for x in metaData['forexWatched']:
			print(x)
			for y in metaData['forexWatched'][x]:
				if self.forexPage['A' + str(curRow)].value != x + '->' + y: #if it changed, needs to move everything down 1
					self.forexPage.move_range('A' + str(curRow) + ':' + get_column_letter(self.forexPage.max_column) + str(self.forexPage.max_row), rows = -1, cols = 0)
				else:
					self.forexPage.cell(row=curRow, column=1, value=x + '->' + y)
					curRow += 1